#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析“全校课表（无自习）xlsx”并输出统一结构。

优先使用 openpyxl（要求环境已安装），保留 OOXML 兜底解析。

目标：从“修改版2025-2026年学年高中课表2026年2月.xlsx”这类全校课表中
提取班级课表，供“带自习年级课表缺失目标班级/教师”时回退使用。

输出结构：
{
  "source_file": "...",
  "parsed_at": "...",
  "schedule_kind": "school_schedule_no_selfstudy",
  "parser": "openpyxl|ooxml",
  "sheets": [
    {
      "name": "Sheet1",
      "classes": [
        {
          "class": "2504班",
          "schedule": {
            "1": {"一": "语文", "二": "数学", ...},
            ...
          }
        }
      ]
    }
  ]
}

注意：严格不猜测，只抽取明确可识别单元格。
"""

from __future__ import annotations

import argparse
import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
from zipfile import ZipFile

VALID_SLOTS = {"早", "1", "2", "3", "4", "5", "6", "7", "活"}
DAY_ORDER = ["一", "二", "三", "四", "五"]
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def qn(name: str) -> str:
    return f"{{{NS_MAIN}}}{name}"


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        if ch.isalpha():
            n = n * 26 + ord(ch.upper()) - 64
    return n


def split_ref(ref: str) -> Tuple[str, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", ref)
    if not m:
        raise ValueError(f"非法单元格引用: {ref}")
    return m.group(1), int(m.group(2))


def parse_shared_strings(zf: ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []

    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    out: List[str] = []
    for si in root.findall(qn("si")):
        texts = []
        for t in si.iterfind(f".//{qn('t')}"):
            texts.append(t.text or "")
        out.append("".join(texts))
    return out


def cell_value(cell: ET.Element, shared: List[str]) -> str:
    ctype = cell.attrib.get("t")
    if ctype == "s":
        v = cell.find(qn("v"))
        if v is None or v.text is None:
            return ""
        idx = int(v.text)
        return shared[idx] if 0 <= idx < len(shared) else ""

    if ctype == "inlineStr":
        texts = []
        for t in cell.iterfind(f".//{qn('t')}"):
            texts.append(t.text or "")
        return "".join(texts)

    v = cell.find(qn("v"))
    return v.text if (v is not None and v.text is not None) else ""


def load_workbook_ooxml(path: Path) -> List[Dict]:
    with ZipFile(path) as zf:
        shared = parse_shared_strings(zf)

        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels
            if rel.tag.endswith("Relationship") and "Id" in rel.attrib and "Target" in rel.attrib
        }

        out_sheets: List[Dict] = []
        sheets_node = wb.find(qn("sheets"))
        if sheets_node is None:
            return out_sheets

        for sh in sheets_node:
            name = sh.attrib.get("name", "")
            rid = sh.attrib.get(f"{{{NS_REL}}}id", "")
            target = rid_to_target.get(rid, "")
            if not target:
                continue
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")

            root = ET.fromstring(zf.read(target))
            sheet_data = root.find(qn("sheetData"))
            data: Dict[Tuple[int, int], str] = {}
            max_row = 0
            max_col = 0

            if sheet_data is not None:
                for row in sheet_data.findall(qn("row")):
                    r = int(row.attrib.get("r", "0"))
                    max_row = max(max_row, r)
                    for cell in row.findall(qn("c")):
                        ref = cell.attrib.get("r")
                        if not ref:
                            continue
                        col_letters, rr = split_ref(ref)
                        c = col_to_num(col_letters)
                        val = norm(cell_value(cell, shared))
                        if val != "":
                            data[(rr, c)] = val
                        max_row = max(max_row, rr)
                        max_col = max(max_col, c)

            out_sheets.append(
                {
                    "title": name,
                    "data": data,
                    "max_row": max_row,
                    "max_col": max_col,
                }
            )

        return out_sheets


def load_workbook_openpyxl(path: Path) -> List[Dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    out_sheets: List[Dict] = []

    for ws in wb.worksheets:
        data: Dict[Tuple[int, int], str] = {}
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = norm(ws.cell(r, c).value)
                if v != "":
                    data[(r, c)] = v

        out_sheets.append(
            {
                "title": ws.title,
                "data": data,
                "max_row": max_row,
                "max_col": max_col,
            }
        )

    return out_sheets


def parse_sheet_by_blocks(sheet: Dict) -> Dict:
    """按“班主任行 + 星期行 + 节次行”块解析。"""
    max_r, max_c = sheet["max_row"], sheet["max_col"]
    data = sheet["data"]

    def cell(r: int, c: int) -> str:
        return norm(data.get((r, c), ""))

    # 1) 找班级标题行：同一行至少出现 2 个“xxxx班”
    block_headers: List[int] = []
    for r in range(1, max_r + 1):
        classes_in_row = 0
        for c in range(1, max_c + 1):
            v = cell(r, c)
            if re.search(r"\d{4}班", v):
                classes_in_row += 1
        if classes_in_row >= 2:
            block_headers.append(r)

    out_classes = {}

    for hr in block_headers:
        day_row = hr + 1
        if day_row > max_r:
            continue

        # 2) 当前块的班级起始列（从标题行提取）
        class_starts = []
        for c in range(1, max_c + 1):
            v = cell(hr, c)
            m = re.search(r"(\d{4})班", v)
            if m:
                class_starts.append((c, f"{m.group(1)}班"))

        if not class_starts:
            continue

        # 3) 星期列映射（day_row）
        day_cols = {c: cell(day_row, c) for c in range(1, max_c + 1)}

        # 为每个班提取其 5 天列：取“班起始列之后最近的 一~五 各一列”
        for class_col, class_name in class_starts:
            picked = {}
            # 搜索窗口适当放大，覆盖后半区块
            for c in range(class_col + 1, min(max_c, class_col + 18) + 1):
                d = day_cols.get(c, "")
                if d in DAY_ORDER and d not in picked:
                    picked[d] = c
                if len(picked) == 5:
                    break

            if len(picked) < 5:
                # 本班星期列不完整，跳过（避免猜测）
                continue

            schedule = out_classes.setdefault(class_name, {})

            # 4) 逐行读取节次
            for r in range(day_row + 1, max_r + 1):
                slot = cell(r, class_col)

                # 到下一块时会出现“高一/高二/高三”或新的班主任行，跳出
                if slot and ("高" in slot and "年" in slot):
                    break
                if re.search(r"\d{4}班", slot):
                    break

                if slot not in VALID_SLOTS:
                    continue

                if slot not in schedule:
                    schedule[slot] = {}

                for d in DAY_ORDER:
                    c = picked[d]
                    schedule[slot][d] = cell(r, c)

    classes = []
    for cls, sched in sorted(out_classes.items()):
        if sched:
            classes.append({"class": cls, "schedule": sched})

    return {"name": sheet["title"], "classes": classes}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument(
        "--parser",
        choices=["auto", "openpyxl", "ooxml"],
        default="auto",
        help="解析引擎：auto(默认优先openpyxl)、openpyxl、ooxml",
    )
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    parser_used = ""
    sheets: List[Dict] = []

    if args.parser in ("auto", "openpyxl"):
        try:
            sheets = load_workbook_openpyxl(in_path)
            parser_used = "openpyxl"
        except Exception:
            if args.parser == "openpyxl":
                raise

    if not sheets:
        sheets = load_workbook_ooxml(in_path)
        parser_used = "ooxml"

    parsed = {
        "source_file": str(in_path),
        "parsed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "schedule_kind": "school_schedule_no_selfstudy",
        "parser": parser_used,
        "sheet_count": len(sheets),
        "sheets": [],
    }

    for sheet in sheets:
        parsed["sheets"].append(parse_sheet_by_blocks(sheet))

    out_path.write_text(json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8")

    # summary
    classes = set()
    for s in parsed["sheets"]:
        for c in s.get("classes", []):
            classes.add(c["class"])

    print(f"[OK] 输出: {out_path}")
    print(f"[INFO] 解析引擎: {parser_used}")
    print(f"[INFO] 班级数: {len(classes)}")
    if classes:
        print(f"[INFO] 班级: {', '.join(sorted(classes))}")


if __name__ == "__main__":
    main()
